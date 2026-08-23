#!/usr/bin/env python3
"""Audit an ELAN corpus against the annotation-information spreadsheet.

The corpus holds most recordings several times over, and the copies disagree:
different Word tier variants, different spellings of the same signer, and
sometimes an age or a gender letter that contradicts the recording sheet. This
script reports all of it and changes nothing.

    python3 audit_corpus.py CORPUS_ROOT \
        --sheet "CORPUS_ROOT/進捗状況-Annotation information.xlsx" \
        -output_folder corpus_audit

What it writes
--------------
    duplicate_recordings.csv       recordings existing in more than one place,
                                   one row per copy, with every Word tier and
                                   its annotation count, and every spelling of
                                   a signer ID inside
    word_variants_by_file.csv      which files carry a Word-jp variant, which
                                   variant, and how many annotations on it
    signer_label_inflation.csv     every signer written more than one way on a
                                   Word tier: one row per tier, with a stable
                                   label number, the tier and the file it came
                                   from
    signer_id_inconsistencies.csv  every way a signer ID goes wrong: several
                                   spellings in one file, a signer from another
                                   recording, a tier name disagreeing with its
                                   PARTICIPANT attribute, and an age or gender
                                   contradicting the spreadsheet


The spreadsheet's first sheet is the authority for who each signer is. It is
laid out as one block per prefecture, two rows per pair, with the age given
once when both signers share it.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# --------------------------------------------------------------------------
# Conventions of this corpus
# --------------------------------------------------------------------------

#: A Word tier is any tier whose ID contains this, case-insensitive.
WORD_TIER_MARKER = "word-jp"

#: ...except a romanised one, which is a transcription, not an annotation.
ROMAN_MARKER = "roman"

#: Folders whose name starts with this are superseded copies.
OLD_FOLDER_PREFIX = "old"

#: Sheet 1 names prefectures in Japanese and English; the corpus uses codes.
PREFECTURE_CODES = {
    "群馬": "GM", "Gunma": "GM",
    "奈良": "NR", "Nara": "NR",
    "長崎": "NS", "Nagasaki": "NS",
    "福岡": "FO", "Fukuoka": "FO",
    "石川": "IS", "Ishikawa": "IS",
    "富山": "TY", "Toyama": "TY",
    "茨城": "IK", "Ibaraki": "IK",
}

#: ``FO_03`` at the front of a participant ID or tier name.
SIGNER_STEM = re.compile(r"^([A-Za-z]{2})[ _-]?0*(\d+)")

#: ``40F`` at the end of the descriptive part: age band then gender letter.
AGE_GENDER = re.compile(r"(?<!\d)(\d{2})\s*([MFmf])(?![A-Za-z0-9])")


def signer_stem(text: str) -> str:
    """``FO_03_NG_40F`` -> ``FO_03``; anything unrecognised comes back whole."""
    match = SIGNER_STEM.match(str(text or "").strip())
    if not match:
        return str(text or "").strip()
    return f"{match.group(1).upper()}_{int(match.group(2)):02d}"


def participant_from_tier_id(tier_id: str) -> str:
    """``FO_03_NG_40F-Word-jp`` -> ``FO_03_NG_40F``."""
    return re.split(r"[-_]?Word[-_ ]?jp", str(tier_id or ""),
                    flags=re.IGNORECASE)[0].strip("-_ ")


def word_variant(tier_id: str) -> Optional[str]:
    """What follows ``Word-jp`` in a tier name; ``""`` for the plain tier.

    Returns None when this is not a Word tier at all.
    """
    match = re.search(r"Word[-_ ]?jp(.*)$", str(tier_id or ""), flags=re.IGNORECASE)
    if not match:
        return None
    return match.group(1).strip()


def local_name(tag: str) -> str:
    return tag.split("}", 1)[-1] if "}" in tag else tag


# --------------------------------------------------------------------------
# The spreadsheet: who each signer actually is
# --------------------------------------------------------------------------

def read_signer_sheet(path: Path, sheet: Optional[str] = None) -> Dict[str, Dict[str, str]]:
    """``{"FO_09": {"age": "40", "gender": "M", "hand": "R", ...}}``.

    Age is written once for a pair when both signers share it, so it carries
    down to the second row; gender and dominant hand are always per signer.
    """
    try:
        import openpyxl
    except ImportError:                                        # pragma: no cover
        print("ERROR: openpyxl is needed to read the spreadsheet.\n"
              "       pip install openpyxl", file=sys.stderr)
        raise SystemExit(2)

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = book[sheet] if sheet else book[book.sheetnames[0]]

    signers: Dict[str, Dict[str, str]] = {}
    code = ""
    age = ""

    for row in worksheet.iter_rows(values_only=True):
        cells = ["" if value is None else str(value).strip() for value in row[:6]]
        heading, pair, number, hand, age_cell, gender = (cells + [""] * 6)[:6]

        if heading:                       # a prefecture block starts here
            for name, prefix in PREFECTURE_CODES.items():
                if name in heading:
                    code, age = prefix, ""
                    break
            continue

        if not code or not number.strip().isdigit():
            continue

        if pair.strip():                  # first row of a pair resets the age
            age = age_cell.strip() or ""
        elif age_cell.strip():            # second row may state its own
            age = age_cell.strip()

        signer = f"{code}_{int(number):02d}"
        signers[signer] = {
            "signer": signer,
            "prefecture": code,
            "pair": pair.strip(),
            "age": (age_cell.strip() or age),
            "gender": gender.strip().upper(),
            "dominant_hand": hand.strip().upper(),
        }

    return signers


# --------------------------------------------------------------------------
# The ELAN files
# --------------------------------------------------------------------------

class FileReport:
    """What one .eaf holds, as far as this audit cares."""

    def __init__(self, path: Path, root: Path):
        self.path = path
        self.relative = path.relative_to(root)
        self.recording = re.sub(r"\.eaf(\.\d+)?$", "", path.name)
        self.error: Optional[str] = None
        #: tier_id -> (participant, variant, n_annotations)
        self.word_tiers: Dict[str, Tuple[str, str, int]] = {}
        #: every participant string seen on any tier, Word or not
        self.participants: Dict[str, List[str]] = defaultdict(list)
        #: the subset written on Word tiers -- these are the ones that end up
        #: counted as signers by the pipelines
        self.word_participants: Dict[str, List[str]] = defaultdict(list)
        #: Word tier ID -> its PARTICIPANT attribute, to compare with the name
        self.tier_attribute: Dict[str, str] = {}

    @property
    def annotations(self) -> int:
        """Every Word tier, plain and variant. Rarely the number you want."""
        return sum(item[2] for item in self.word_tiers.values())

    @property
    def plain_annotations(self) -> int:
        """The ``-Word-jp`` tiers only -- the sanctioned gloss."""
        return sum(count for _, variant, count in self.word_tiers.values()
                   if not variant)

    @property
    def variant_annotations(self) -> int:
        return sum(count for _, variant, count in self.word_tiers.values()
                   if variant)

    @property
    def variants(self) -> List[str]:
        return sorted({item[1] for item in self.word_tiers.values() if item[1]})

    @property
    def signers(self) -> List[str]:
        """Signers named on the Word tiers -- the ones the pipelines read.

        Deliberately not every tier: a file also carries gesture, mouth and
        comment tiers with naming habits of their own, and a signer problem
        there is not a problem with the gloss.
        """
        return sorted({signer_stem(name) for name in self.word_participants})

    def signer_variants(self) -> Dict[str, List[str]]:
        """Signers written more than one way on the Word tiers of this file.

        Word tiers only: those are what the pipelines read, so those are the
        spellings that turn one person into two signers. Comment and gesture
        tiers have their own naming habits and are not worth reporting here.
        """
        grouped: Dict[str, set] = defaultdict(set)
        for name in self.word_participants:
            grouped[signer_stem(name)].add(name)
        return {signer: sorted(names)
                for signer, names in grouped.items() if len(names) > 1}


def read_eaf(path: Path, root: Path) -> FileReport:
    report = FileReport(path, root)
    try:
        tree = ET.parse(path)
    except ET.ParseError as error:
        report.error = f"unreadable XML: {error}"
        return report

    for tier in tree.getroot().iter():
        if local_name(tier.tag) != "TIER":
            continue

        tier_id = tier.attrib.get("TIER_ID", "")
        variant = word_variant(tier_id)
        is_word = variant is not None and ROMAN_MARKER not in tier_id.lower()

        attribute = tier.attrib.get("PARTICIPANT", "").strip()
        if attribute:
            report.participants[attribute].append(tier_id)

        if not is_word:
            # Outside a Word tier the name is the tier's own
            # (``FO_01_KT_70F-GS-en``); reading it as a participant would
            # invent a signer per tier. Only the attribute counts there.
            continue

        # On a Word tier the layout ``<participant>-Word-jp`` is reliable, and
        # it is what the parser reads -- so it, not the PARTICIPANT attribute,
        # decides who the pipelines think signed this. The two disagree more
        # often than one would like, which is its own report.
        from_name = participant_from_tier_id(tier_id) or attribute
        if from_name:
            report.participants[from_name].append(tier_id)
            report.word_participants[from_name].append(tier_id)
        report.tier_attribute[tier_id] = attribute

        filled = sum(
            1 for element in tier.iter()
            if local_name(element.tag) == "ANNOTATION"
            and any((child.text or "").strip()
                    for child in element.iter()
                    if local_name(child.tag) == "ANNOTATION_VALUE")
        )
        report.word_tiers[tier_id] = (
            participant_from_tier_id(tier_id) or participant, variant, filled
        )

    return report


def find_eaf_files(root: Path, include_old: bool) -> List[Path]:
    found: List[Path] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        if not re.search(r"\.eaf(\.\d+)?$", path.name):
            continue
        if not include_old and any(part.lower().startswith(OLD_FOLDER_PREFIX)
                                   for part in path.relative_to(root).parts[:-1]):
            continue
        found.append(path)
    return found


# --------------------------------------------------------------------------
# Reports
# --------------------------------------------------------------------------

def write_csv(path: Path, rows: List[dict], columns: List[str]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)
    return path


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="audit_corpus.py",
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("corpus_root", type=Path,
                        help="Folder to walk, e.g. ELAN_W_OpenPose.")
    parser.add_argument("--sheet-file", "--sheet", dest="sheet_file", type=Path,
                        default=None, metavar="XLSX",
                        help="Annotation-information workbook. Without it the "
                             "metadata comparison is skipped.")
    parser.add_argument("--sheet-name", dest="sheet_name", default=None,
                        help="Worksheet to read (default: the first one).")
    parser.add_argument("-output_folder", "--output-folder", dest="output_folder",
                        type=Path, default=Path("corpus_audit"))
    parser.add_argument("--include-old", action="store_true",
                        help="Also walk folders whose name begins with 'old'.")
    return parser


def main(argv=None) -> int:
    args = build_argument_parser().parse_args(argv)

    root = args.corpus_root.expanduser().resolve()
    if not root.is_dir():
        print(f"ERROR: not a folder: {root}", file=sys.stderr)
        return 1
    output = args.output_folder.expanduser().resolve()

    sheet: Dict[str, Dict[str, str]] = {}
    if args.sheet_file:
        sheet = read_signer_sheet(args.sheet_file.expanduser().resolve(),
                                  args.sheet_name)

    paths = find_eaf_files(root, args.include_old)
    print(f"Corpus:  {root}")
    print(f"Files:   {len(paths)}")
    print(f"Sheet:   {len(sheet)} signers" if sheet else "Sheet:   (not given)")
    print()

    reports = [read_eaf(path, root) for path in paths]
    by_recording: Dict[str, List[FileReport]] = defaultdict(list)
    for report in reports:
        by_recording[report.recording].append(report)

    # ---- recordings that exist more than once ---------------------------
    # One row per copy, spelling out every Word tier it holds and every way a
    # signer is written in it. Two copies of one recording routinely differ in
    # both, and a total alone hides that: it is the tier list that shows a
    # second pass named plainly in one copy and ``-T`` in another.
    duplicate_rows = []
    for recording in sorted(by_recording):
        copies = by_recording[recording]
        if len(copies) < 2:
            continue
        for copy in sorted(copies, key=lambda item: (-item.plain_annotations,
                                                     str(item.relative))):
            duplicate_rows.append({
                "recording": recording,
                "n_copies": len(copies),
                "path": str(copy.relative),
                "word_jp_annotations": copy.plain_annotations,
                "word_tiers": " ".join(
                    f"{tier_id}={count}"
                    for tier_id, (_, _, count) in sorted(copy.word_tiers.items())),
                "signer_ids": " ".join(sorted(copy.word_participants)),
                "signers": " ".join(copy.signers),
                "bytes": copy.path.stat().st_size,
            })

    # ---- which files carry a variant, and which variant -----------------
    variant_by_file = []
    for report in reports:
        if not report.variants:
            continue
        counts = defaultdict(int)
        for _, variant, count in report.word_tiers.values():
            if variant:
                counts[variant] += count
        variant_by_file.append({
            "recording": report.recording,
            "path": str(report.relative),
            "variants": " ".join(report.variants),
            "n_variant_tiers": sum(1 for _, variant, _ in report.word_tiers.values()
                                   if variant),
            "annotations_by_variant": " ".join(
                f"{variant}={counts[variant]}" for variant in sorted(counts)),
            "variant_annotations": report.variant_annotations,
            "plain_word_jp_annotations": report.plain_annotations,
        })

    # ---- where a signer's extra labels come from ------------------------
    # The pipelines count distinct participant strings on Word tiers, so a
    # signer written two ways is two signers. One row per place a label is
    # actually written -- signer, label, the tier, the file -- so a name can be
    # traced to the tiers that have to be edited to repair it.
    written_as: Dict[str, set] = defaultdict(set)
    for report in reports:
        for participant, _, _ in report.word_tiers.values():
            written_as[signer_stem(participant)].add(participant)

    # A stable number per label, so one glance down the column shows which
    # spelling a given file used. Alphabetical, so it does not move between
    # runs.
    label_number = {
        signer: {label: index for index, label in enumerate(sorted(names), 1)}
        for signer, names in written_as.items()
    }

    label_rows = []
    for report in reports:
        for tier_id, (participant, _, _) in sorted(report.word_tiers.items()):
            signer = signer_stem(participant)
            if len(written_as[signer]) < 2:
                continue
            label_rows.append({
                "signer": signer,
                "n_labels_for_this_signer": len(written_as[signer]),
                "label_no": label_number[signer][participant],
                "label": participant,
                "tier_id": tier_id,
                "file": str(report.relative),
                "recording": report.recording,
            })
    label_rows.sort(key=lambda row: (row["signer"], row["label_no"],
                                     row["recording"], row["file"]))

    # ---- everything wrong with a signer ID, in one place -----------------
    # Four ways a signer ID goes wrong, gathered into one file with a
    # ``problem`` column, because they are all the same question for whoever
    # has to fix them: who is this tier actually about?
    inconsistent = []

    def flag(report, signer, problem, detail, tiers):
        inconsistent.append({
            "recording": report.recording,
            "path": str(report.relative),
            "signer": signer,
            "problem": problem,
            "detail": detail,
            "tiers": tiers,
        })

    for report in reports:
        # one person, several spellings, in the same file
        for signer, spellings in sorted(report.signer_variants().items()):
            flag(report, signer, "several spellings in one file",
                 " / ".join(spellings),
                 " ".join(sorted(tier for spelling in spellings
                                 for tier in report.word_participants[spelling])))

    # ---- signers who do not belong to this recording --------------------
    # ``NS_07-08_Cur`` should contain NS_07 and NS_08 and nobody else. A tier
    # for someone else is a template copied from another recording and never
    # renamed -- which also means its annotations are attributed to the wrong
    # person.
    for report in reports:
        match = re.match(r"^\d?([A-Za-z]{2})_(\d+)-(\d+)", report.recording)
        if not match:
            continue
        prefix = match.group(1).upper()
        belongs = {f"{prefix}_{int(match.group(2)):02d}",
                   f"{prefix}_{int(match.group(3)):02d}"}
        for signer in report.signers:
            if not re.fullmatch(r"[A-Z]{2}_\d+", signer) or signer in belongs:
                continue
            flag(report, signer, "signer from another recording",
                 f"this recording is {' and '.join(sorted(belongs))}",
                 " ".join(sorted(tier for name, tiers in report.word_participants.items()
                                 if signer_stem(name) == signer for tier in tiers)))

    # ---- Word tiers whose name and PARTICIPANT attribute disagree -------
    for report in reports:
        for tier_id, attribute in sorted(report.tier_attribute.items()):
            from_name = participant_from_tier_id(tier_id)
            if not attribute or not from_name or attribute == from_name:
                continue
            same = signer_stem(from_name) == signer_stem(attribute)
            flag(report, signer_stem(from_name),
                 "tier name and PARTICIPANT disagree"
                 + ("" if same else " (different signer)"),
                 f"name says {from_name}, PARTICIPANT says {attribute}",
                 tier_id)

    # ---- metadata: what the tier claims against what the sheet says -----
    if sheet:
        seen: set = set()
        for report in reports:
            for spelling, tiers in sorted(report.word_participants.items()):
                signer = signer_stem(spelling)
                key = (report.recording, str(report.relative), spelling)
                if key in seen:
                    continue
                seen.add(key)

                truth = sheet.get(signer)
                match = AGE_GENDER.search(spelling)
                if not match:
                    # No age/gender in the name is normal for plain IDs like
                    # "memo"; only report it when it looks like a signer.
                    if truth is not None and spelling != signer:
                        flag(report, signer, "no age/gender in the ID",
                             spelling, " ".join(sorted(tiers)))
                    continue

                if truth is None:
                    flag(report, signer, "not in the spreadsheet",
                         spelling, " ".join(sorted(tiers)))
                    continue

                in_tier_age, in_tier_gender = match.group(1), match.group(2).upper()
                problems = []
                if truth["age"] and in_tier_age != truth["age"]:
                    problems.append(("age", in_tier_age, truth["age"]))
                if truth["gender"] and in_tier_gender != truth["gender"]:
                    problems.append(("gender", in_tier_gender, truth["gender"]))

                for field, in_tier, in_sheet in problems:
                    flag(report, signer, f"{field} contradicts the spreadsheet",
                         f"{spelling} says {in_tier}, sheet says {in_sheet}",
                         " ".join(sorted(tiers)))

    inconsistent.sort(key=lambda row: (row["problem"], row["signer"],
                                       row["recording"], row["path"]))

    written = [
        write_csv(output / "duplicate_recordings.csv", duplicate_rows,
                  list(duplicate_rows[0].keys()) if duplicate_rows
                  else ["recording", "n_copies", "path", "word_jp_annotations",
                        "word_tiers", "signer_ids"]),
        write_csv(output / "word_variants_by_file.csv", variant_by_file,
                  list(variant_by_file[0].keys()) if variant_by_file
                  else ["recording", "path", "variants"]),
        write_csv(output / "signer_label_inflation.csv", label_rows,
                  list(label_rows[0].keys()) if label_rows
                  else ["signer", "label", "agrees_with_sheet", "annotations"]),
        write_csv(output / "signer_id_inconsistencies.csv", inconsistent,
                  list(inconsistent[0].keys()) if inconsistent
                  else ["recording", "path", "signer", "problem", "detail",
                        "tiers"]),
    ]

    duplicated = {name: items for name, items in by_recording.items() if len(items) > 1}
    print(f"recordings                       : {len(by_recording):5d}")
    print(f"  existing in more than one place: {len(duplicated):5d}"
          f"  ({sum(len(v) for v in duplicated.values()) - len(duplicated)} extra copies)")
    print(f"files carrying a Word-jp variant : {len(variant_by_file):5d}")
    print()
    people = {row["signer"] for row in label_rows}
    all_labels = {(row["signer"], row["label"]) for row in label_rows}
    print(f"signers written more than one way: {len(people):5d}"
          f"  ({len(all_labels)} labels, {len(all_labels) - len(people)} more than"
          f" there are people, over {len(label_rows)} tiers)")
    print(f"signer ID problems               : {len(inconsistent):5d}")
    for problem in sorted({row["problem"] for row in inconsistent}):
        rows = [row for row in inconsistent if row["problem"] == problem]
        print(f"    {problem:42s} {len(rows):5d}"
              f"  ({len({row['signer'] for row in rows})} signers,"
              f" {len({row['path'] for row in rows})} files)")
    unreadable = [r for r in reports if r.error]
    if unreadable:
        print(f"files that would not parse       : {len(unreadable):5d}")

    print()
    for path in written:
        print(f"  {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
